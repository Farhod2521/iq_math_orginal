from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from rest_framework import status
from .models import  Chapter, Topic, Question, Choice, CompositeSubQuestion
from .serializers import(
    SubjectSerializer, MyChapterAddSerializer, MyTopicAddSerializer,
    ChoiceSerializer, CompositeSubQuestionSerializer, QuestionSerializer, SubjectRegisterSerilzier, OpenAIQuestionSerializer
)
from django_app.app_user.models import  Subject
from django.shortcuts import get_object_or_404
from rest_framework.generics import ListAPIView
import os
from dotenv import load_dotenv

class SubjectListAPIView(ListAPIView):
    queryset = Subject.objects.all()
    serializer_class = SubjectRegisterSerilzier





class TeacherSubjectsAPIView(APIView):
    permission_classes = [IsAuthenticated]  # Faqat tizimga kirgan foydalanuvchilar

    def get(self, request):
        teacher = request.user.teacher_profile  # Tizimga kirgan o‘qituvchini olish
        subjects = Subject.objects.filter(teachers=teacher).order_by("order")  # O‘qituvchiga bog‘langan fanlar
        serializer = SubjectSerializer(subjects, many=True)
        return Response(serializer.data)

    def put(self, request, pk):
        teacher = request.user.teacher_profile
        try:
            subject = Subject.objects.get(pk=pk, teachers=teacher)
        except Subject.DoesNotExist:
            return Response({"error": "Bu fan sizga tegishli emas yoki mavjud emas."}, status=status.HTTP_404_NOT_FOUND)

        serializer = SubjectSerializer(subject, data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def delete(self, request, pk):
        teacher = request.user.teacher_profile
        try:
            subject = Subject.objects.get(pk=pk, teachers=teacher)
        except Subject.DoesNotExist:
            return Response({"error": "Bu fan sizga tegishli emas yoki mavjud emas."}, status=status.HTTP_404_NOT_FOUND)

        subject.delete()
        return Response({"message": "Fan o‘chirildi."}, status=status.HTTP_204_NO_CONTENT)
    
class MyChapterAddCreateView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        teacher = request.user.teacher_profile
        subject_id = request.data.get("subject")

        if not subject_id:
            return Response({"error": "Fan ID majburiy!"}, status=status.HTTP_400_BAD_REQUEST)

        try:
            subject = Subject.objects.get(id=subject_id, teachers=teacher)
        except Subject.DoesNotExist:
            return Response({"error": "Siz bu fanga bo'lim qo'sha olmaysiz!"}, status=status.HTTP_403_FORBIDDEN)

        serializer = MyChapterAddSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save(subject=subject)  # yoki kerakli boshqa fieldlar
            return Response({"message": "Chapter yaratildi!", "data": serializer.data}, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class MyTopicAddCreateView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        teacher = request.user.teacher_profile
        chapter_id = request.data.get("chapter")

        if not chapter_id:
            return Response({"error": "Chapter ID majburiy!"}, status=status.HTTP_400_BAD_REQUEST)

        try:
            chapter = Chapter.objects.get(id=chapter_id, subject__teachers=teacher)
        except Chapter.DoesNotExist:
            return Response({"error": "Siz bu bo‘limga mavzu qo‘sha olmaysiz!"}, status=status.HTTP_403_FORBIDDEN)

        serializer = MyTopicAddSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save(chapter=chapter)  # kerakli bo‘lsa `chapter`ni o‘zing kirit
            return Response({"message": "Topic yaratildi!", "data": serializer.data}, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    

class MyQuestionListView(APIView):
    permission_classes = [IsAuthenticated]
    def get(self, request, id):
        try:
            topic = Topic.objects.get(id=id)
        except Topic.DoesNotExist:
            return Response({"error": "Mavzu topilmadi."}, status=status.HTTP_404_NOT_FOUND)
        
        questions = Question.objects.filter(topic=topic).prefetch_related('sub_questions', 'choices')
        serializer = QuestionSerializer(questions, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)
    

class MyChapterListView(APIView):
    """Tizimga kirgan o‘qituvchining barcha bo‘limlarini olish"""
    permission_classes = [IsAuthenticated]

    def get(self, request, id):
        teacher = request.user.teacher_profile  
        subjects = Subject.objects.filter(teachers=teacher)  
        chapters = Chapter.objects.filter(subject__in=[id]).order_by('order') 

        serializer = MyChapterAddSerializer(chapters, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)

    def put(self, request, id):
        chapter = get_object_or_404(Chapter, id=id)
        serializer = MyChapterAddSerializer(chapter, data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def delete(self, request, id):
        chapter = get_object_or_404(Chapter, id=id)
        chapter.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)
    
    
class MyTopicListView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, id):
        topics = Topic.objects.filter(chapter__in=[id]).order_by('order') 
        serializer = MyTopicAddSerializer(topics, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)
    def put(self, request, id):
        topic = get_object_or_404(Topic, id=id)
        serializer = MyTopicAddSerializer(topic, data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def delete(self, request, id):
        topic = get_object_or_404(Topic, id=id)
        topic.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)


################################  QUESTION BY TEACHER CREATE ##################################################
from django.db import transaction
   
class QuestionAddCreateView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        serializer = QuestionSerializer(data=request.data, context={'request': request})
        if serializer.is_valid():
            try:
                with transaction.atomic():
                    question = serializer.save()

                    # Choice savollarini saqlash
                    if question.question_type in ['choice', 'image_choice']:
                        choices_data = request.data.get('choices', [])
                        if not choices_data:
                            raise ValueError("Choice savollar uchun variantlar kiritilmagan")
                        
                        # Avvalgi choicelarni o'chirish
                        question.choices.all().delete()
                        
                        for idx, choice_data in enumerate(choices_data):
                            # Agar letter mavjud bo'lmasa, harflarni avtomatik belgilash
                            if 'letter' not in choice_data or not choice_data['letter']:
                                choice_data['letter'] = chr(65 + idx)  # 65 = 'A' ASCII kodi
                            
                            choice_data['question'] = question.id
                            choice_serializer = ChoiceSerializer(data=choice_data, context={'request': request})
                            if choice_serializer.is_valid():
                                choice = choice_serializer.save()
                                # Rasmni alohida saqlash
                                if 'image' in request.FILES:
                                    image_file = request.FILES.get(f'choices[{idx}].image')
                                    if image_file:
                                        choice.image = image_file
                                        choice.save()
                            else:
                                raise ValueError(f"Variantdagi ma'lumotlar noto‘g‘ri: {choice_serializer.errors}")

                    # Composite savollarini saqlash
                    elif question.question_type == 'composite':
                        sub_questions_data = request.data.get('sub_questions', [])
                        if not sub_questions_data:
                            raise ValueError("Composite savol uchun kichik savollar kiritilmagan")
                        
                        # Avvalgi sub_questionlarni o'chirish
                        question.sub_questions.all().delete()
                        
                        for sub_data in sub_questions_data:
                            sub_data['question'] = question.id
                            sub_serializer = CompositeSubQuestionSerializer(data=sub_data)
                            if sub_serializer.is_valid():
                                sub_serializer.save()
                            else:
                                raise ValueError(f"Kichik savoldagi ma'lumotlar noto‘g‘ri: {sub_serializer.errors}")

                return Response(QuestionSerializer(question, context={'request': request}).data, status=status.HTTP_201_CREATED)
            except Exception as e:
                return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)

        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
################################  QUESTION BY TEACHER DELETE ##################################################

class QuestionDeleteView(APIView):
    permission_classes = [IsAuthenticated]

    def delete(self, request, pk):
        question = get_object_or_404(Question, pk=pk)
        question.delete()
        return Response({"message": "Savol muvaffaqiyatli o‘chirildi."}, status=status.HTTP_204_NO_CONTENT)

################################  QUESTION BY TEACHER UPDATE ##################################################

class QuestionUpdateView(APIView):
    permission_classes = [IsAuthenticated]

    def put(self, request, pk):
        question = get_object_or_404(Question, pk=pk)
        serializer = QuestionSerializer(question, data=request.data, partial=True, context={'request': request})
        
        if serializer.is_valid():
            try:
                with transaction.atomic():
                    question = serializer.save()

                    # Choice savollarini yangilash
                    if question.question_type in ['choice', 'image_choice']:
                        choices_data = request.data.get('choices', [])
                        if not choices_data:
                            raise ValueError("Choice savollar uchun variantlar kiritilmagan")

                        question.choices.all().delete()

                        for idx, choice_data in enumerate(choices_data):
                            if 'letter' not in choice_data or not choice_data['letter']:
                                choice_data['letter'] = chr(65 + idx)

                            choice_data['question'] = question.id
                            choice_serializer = ChoiceSerializer(data=choice_data, context={'request': request})
                            if choice_serializer.is_valid():
                                choice = choice_serializer.save()
                                image_file = request.FILES.get(f'choices[{idx}].image')
                                if image_file:
                                    choice.image = image_file
                                    choice.save()
                            else:
                                raise ValueError(f"Variantdagi ma'lumotlar noto‘g‘ri: {choice_serializer.errors}")

                    # Composite savollarini yangilash
                    elif question.question_type == 'composite':
                        sub_questions_data = request.data.get('sub_questions', [])
                        if not sub_questions_data:
                            raise ValueError("Composite savol uchun kichik savollar kiritilmagan")

                        question.sub_questions.all().delete()

                        for sub_data in sub_questions_data:
                            sub_data['question'] = question.id
                            sub_serializer = CompositeSubQuestionSerializer(data=sub_data)
                            if sub_serializer.is_valid():
                                sub_serializer.save()
                            else:
                                raise ValueError(f"Kichik savoldagi ma'lumotlar noto‘g‘ri: {sub_serializer.errors}")

                return Response(QuestionSerializer(question, context={'request': request}).data, status=status.HTTP_200_OK)
            except Exception as e:
                return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)

        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


from openpyxl import load_workbook
from io import BytesIO
from django.http import HttpResponse
import os
import pandas as pd
class TextQuestionToXlsxImport(APIView):
    def post(self, request):
        question_type = request.data.get("question_type")

        template_path = '/home/user/backend/iq_math_orginal/shablon.xlsx'
        if not os.path.exists(template_path):
            return Response({"error": "shablon.xlsx topilmadi"}, status=status.HTTP_404_NOT_FOUND)

        workbook = load_workbook(template_path)
        sheet = workbook.active

        next_row = sheet.max_row + 1

        sheet.cell(row=next_row, column=1, value=question_type)

        # BytesIO bilan faylni virtual tarzda saqlash
        output = BytesIO()
        workbook.save(output)
        output.seek(0)

        response = HttpResponse(
            output,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=shablon_yangilangan.xlsx'
        return response
    
class ChoiceQuestionToXlsxImport(APIView):
    def post(self, request):
        question_type = request.data.get("question_type")

        template_path = '/home/user/backend/iq_math_orginal/choice.xlsx'
        if not os.path.exists(template_path):
            return Response({"error": "choice.xlsx topilmadi"}, status=status.HTTP_404_NOT_FOUND)

        workbook = load_workbook(template_path)
        sheet = workbook.active

        next_row = sheet.max_row + 1
        sheet.cell(row=next_row, column=1, value=question_type)

        # BytesIO bilan faylni virtual tarzda saqlash
        output = BytesIO()
        workbook.save(output)
        output.seek(0)

        response = HttpResponse(
            output,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=shablon_yangilangan.xlsx'
        return response
    
class CompenQuestionToXlsxImport(APIView):
    def post(self, request):
        question_type = request.data.get("question_type")

        template_path = '/home/user/backend/iq_math_orginal/composite.xlsx'
        if not os.path.exists(template_path):
            return Response({"error": "composite.xlsx topilmadi"}, status=status.HTTP_404_NOT_FOUND)

        workbook = load_workbook(template_path)
        sheet = workbook.active

        next_row = sheet.max_row + 1
        sheet.cell(row=next_row, column=1, value=question_type)

        # BytesIO bilan faylni virtual tarzda saqlash
        output = BytesIO()
        workbook.save(output)
        output.seek(0)

        response = HttpResponse(
            output,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=shablon_yangilangan.xlsx'
        return response 

# class QuestionImportFromXlsx(APIView):
#     def post(self, request):
#         file = request.FILES.get('file')

#         if not file:
#             return Response({"error": "Fayl yuborilmadi"}, status=status.HTTP_400_BAD_REQUEST)

#         try:
#             df = pd.read_excel(BytesIO(file.read()))
#         except Exception as e:
#             return Response({"error": f"Excel faylni o'qishda xatolik: {str(e)}"}, status=status.HTTP_400_BAD_REQUEST)

#         created_count = 0
#         skipped = []

#         for index, row in df.iterrows():
#             topic_id = row.get('topic_id')
#             if not topic_id:
#                 skipped.append(f"{index + 2}-qator: topic_id yo'q")
#                 continue

#             try:
#                 topic = Topic.objects.get(id=topic_id)
#             except Topic.DoesNotExist:
#                 skipped.append(f"{index + 2}-qator: Topic ID {topic_id} topilmadi")
#                 continue

#             question = Question.objects.create(
#                 topic=topic,
#                 question_type='text',  # Agar boshqa type bo‘lsa, excelga qo‘shib o‘zgartirish mumkin
#                 level=row.get('level') or 1,
#                 question_text_uz=row.get('question_text_uz', ''),
#                 question_text_ru=row.get('question_text_ru', ''),
#                 correct_text_answer_uz=row.get('answer_text_uz', ''),
#                 correct_text_answer_ru=row.get('answer_text_ru', ''),
#                 video_url_uz=row.get('video_url_uz', ''),
#                 video_url_ru=row.get('video_url_ru', '')
#             )

#             created_count += 1

#         return Response({
#             "success": True,
#             "created": created_count,
#             "skipped": skipped
#         }, status=status.HTTP_201_CREATED)
    


class QuestionImportFromXlsx(APIView):
    def post(self, request):
        file = request.FILES.get('file')
        topic_id = request.data.get('topic_id')

        if not file:
            return Response({"error": "Fayl yuborilmadi"}, status=status.HTTP_400_BAD_REQUEST)

        if not topic_id:
            return Response({"error": "Topic ID yuborilmadi"}, status=status.HTTP_400_BAD_REQUEST)

        try:
            topic = Topic.objects.get(id=topic_id)
        except Topic.DoesNotExist:
            return Response({"error": f"Topic ID {topic_id} topilmadi"}, status=status.HTTP_400_BAD_REQUEST)

        try:
            df = pd.read_excel(BytesIO(file.read()))
        except Exception as e:
            return Response({"error": f"Excel faylni o'qishda xatolik: {str(e)}"}, status=status.HTTP_400_BAD_REQUEST)

        created_count = 0
        skipped = []

        for index, row in df.iterrows():
            question_type = row.get('question_type')

            if not question_type:
                skipped.append(f"{index + 2}-qator: question_type yo'q")
                continue

            # Asosiy savolni yaratish
            question = Question.objects.create(
                topic=topic,
                question_type=question_type,
                level=row.get('level') or 1,
                question_text_uz=row.get('question_text_uz', ''),
                question_text_ru=row.get('question_text_ru', ''),
                correct_text_answer_uz=row.get('answer_text_uz', ''),
                correct_text_answer_ru=row.get('answer_text_ru', ''),
                video_url_uz=row.get('video_url_uz', ''),
                video_url_ru=row.get('video_url_ru', '')
            )

            if question_type == 'choice':
                # Choice variantlarini yaratish
                Choice.objects.create(
                    question=question,
                    letter=row.get('letter', ''),
                    text_uz=row.get('text_uz', ''),
                    text_ru=row.get('text_ru', ''),
                    is_correct=row.get('is_correct', False)
                )

            elif question_type == 'composite':
                # CompositeSubQuestion variantlarini yaratish
                CompositeSubQuestion.objects.create(
                    question=question,
                    text1_uz=row.get('text1_uz', ''),
                    text1_ru=row.get('text1_ru', ''),
                    correct_answer_uz=row.get('correct_answer_uz', ''),
                    correct_answer_ru=row.get('correct_answer_ru', ''),
                    text2_uz=row.get('text2_uz', ''),
                    text2_ru=row.get('text2_ru', '')
                )

            # Agar boshqa turdagi savol bo'lsa (masalan, text), hech narsa qilmaymiz
            created_count += 1

        return Response({
            "success": True,
            "created": created_count,
            "skipped": skipped
        }, status=status.HTTP_201_CREATED)
    
import re
from rest_framework.parsers import MultiPartParser
class UploadQuestionsAPIView(APIView):
    parser_classes = [MultiPartParser]

    def post(self, request, *args, **kwargs):
        topic_id = request.data.get('topic_id')
        question_type = request.data.get('question_type')
        level = request.data.get('level')
        file = request.data.get('file')

        if not all([topic_id, question_type, level, file]):
            return Response({"error": "Missing fields"}, status=status.HTTP_400_BAD_REQUEST)

        try:
            topic = Topic.objects.get(id=topic_id)
        except Topic.DoesNotExist:
            return Response({"error": "Topic not found"}, status=status.HTTP_404_NOT_FOUND)

        content = file.read().decode('utf-8')

        try:
            # 1. Hamma bloklarni topib olamiz (to'rt qismli: uz_savol, uz_javob, ru_savol, ru_javob)
            pattern = re.compile(
                r'\\begin\{minipage\}.*?\\centering\s*(.*?)\\begin\{enumerate\}(.*?)\\end\{enumerate\}.*?'
                r'\\begin\{enumerate\}(.*?)\\end\{enumerate\}.*?'
                r'\\begin\{minipage\}.*?\\centering\s*\\begin\{enumerate\}.*?\\item\s*(.*?)\\end\{enumerate\}.*?'
                r'\\begin\{enumerate\}(.*?)\\end\{enumerate\}',
                re.DOTALL
            )

            matches = pattern.findall(content)

            if not matches:
                return Response({"error": "No questions found"}, status=status.HTTP_400_BAD_REQUEST)

            for match in matches:
                uz_question_title = match[0].strip()
                uz_question_body = match[1].strip()
                uz_answers = match[2].strip()
                ru_question_title = match[3].strip()
                ru_answers = match[4].strip()

                # Har bir `item` (ya'ni bitta savol) bo'yicha ajratamiz
                uz_savollar = re.findall(r'\\item\s*(.+)', uz_question_body)
                uz_javoblar = re.findall(r'\\item\s*(.+)', uz_answers)
                ru_savollar = re.findall(r'\\item\s*(.+)', ru_question_title)  # Eslatma: ru_savollar aslida bitta "Vychislite..." degan sarlavha bo'ladi
                ru_javoblar = re.findall(r'\\item\s*(.+)', ru_answers)

                # Agar savol ru_savollar topilmasa, umumiy sarlavhani savol sifatida olamiz
                ru_question_title_text = ru_savollar[0] if ru_savollar else ''

                for i in range(len(uz_savollar)):
                    question = Question(
                        topic=topic,
                        question_text_uz=uz_savollar[i].strip(),
                        correct_text_answer_uz=uz_javoblar[i].strip() if i < len(uz_javoblar) else '',
                        question_text_ru=uz_savollar[i].strip(),  # O'zbekcha savolni ruschasiga vaqtincha o'zbekcha yozamiz
                        correct_text_answer_ru=ru_javoblar[i].strip() if i < len(ru_javoblar) else '',
                        question_type=question_type,
                        level=level,
                    )
                    question.save()

            return Response({"message": "Savollar muvaffaqiyatli yuklandi"}, status=status.HTTP_201_CREATED)

        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        











from rest_framework.pagination import PageNumberPagination



class OpenAIStandardResultsSetPagination(PageNumberPagination):
    page_size = 10
    page_size_query_param = 'page_size'
    max_page_size = 100

class OpenAIQuestionListView(ListAPIView):
    queryset = Question.objects.all().order_by('id')
    serializer_class = OpenAIQuestionSerializer
    pagination_class = OpenAIStandardResultsSetPagination


from openai import OpenAI
import json
import base64
from bs4 import BeautifulSoup
from django.conf import settings
client = OpenAI(
    api_key=f"{os.getenv('OPENAI')}"
)

# class OpenAIProcessAPIView(APIView):
#     def post(self, request, *args, **kwargs):
#         input_text = request.data.get('text', '').strip()

#         if not input_text:
#             return Response({'error': 'Matn kiritilmagan'}, status=status.HTTP_400_BAD_REQUEST)

#         try:
#             completion = client.chat.completions.create(
#                 model="gpt-4o-mini",
#                 messages=[
#                     {
#                         "role": "user",
#                         "content": "Savol ishlanish yo'li bilan ishlab ber ketma ketlikda qisqa lo'nda aniq javob ber: " + input_text
#                     }
#                 ]
#             )
#         except Exception as e:
#             return Response({'error': f'AI bilan bog‘lanishda xatolik: {str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

#         result_content = completion.choices[0].message.content

#         # Natijani JSON formatda qaytarishga harakat qilamiz, agar bo'lmasa matn sifatida qaytariladi
#         try:
#             result_data = json.loads(result_content)
#             return Response({'result': result_data})
#         except json.JSONDecodeError:
#             return Response({'result': result_content})

def save_base64_image(base64_string, extension, filename):
    directory = os.path.join(settings.MEDIA_ROOT, 'question_images')
    os.makedirs(directory, exist_ok=True)
    image_data = base64.b64decode(base64_string)
    filepath = os.path.join(directory, f"{filename}.{extension}")
    with open(filepath, "wb") as f:
        f.write(image_data)
    return settings.MEDIA_URL + f"question_images/{filename}.{extension}"

# HTMLdan matn va base64 rasmlarni ajratish
def clean_html_and_extract_images(html, question_id):
    soup = BeautifulSoup(html, "html.parser")
    image_urls = []

    for i, img in enumerate(soup.find_all("img")):
        src = img.get("src", "")
        match = re.match(r'data:image/(png|jpeg);base64,(.*)', src)
        if match:
            ext = match.group(1)
            base64_data = match.group(2)
            filename = f"question_{question_id}_img{i+1}"
            image_url = save_base64_image(base64_data, ext, filename)
            image_urls.append(f"[Rasm {i+1}]: {image_url}")
            img.replace_with(f"[Rasm {i+1}]")  # HTML dan o‘rniga matn qo‘yiladi

    clean_text = soup.get_text(separator=" ", strip=True)
    return clean_text, image_urls

# OpenAI API uchun view
class OpenAIProcessAPIView(APIView):
    def post(self, request, *args, **kwargs):
        question_id = request.data.get('question_id')
        lang = request.data.get('lang')
        question_type = request.data.get('question_type')

        if not question_id or not lang or not question_type:
            return Response({'error': 'Kerakli parametrlar yuborilmagan'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            question = Question.objects.get(id=question_id)
        except Question.DoesNotExist:
            return Response({'error': 'Savol topilmadi'}, status=status.HTTP_404_NOT_FOUND)

        # Tilga qarab savol matnini olish
        if lang == 'uz':
            html = question.question_text_uz
        elif lang == 'ru':
            html = question.question_text_ru
        else:
            return Response({'error': 'Til noto‘g‘ri ko‘rsatilgan (faqat uz yoki ru)'}, status=400)

        soup = BeautifulSoup(html, "html.parser")
        question_text = soup.get_text()

        # Rasm URL'larini topish
        image_urls = []
        for img_tag in soup.find_all("img"):
            src = img_tag.get("src", "")
            if not src:
                continue
            full_url = src if src.startswith("http") else request.build_absolute_uri(src)
            image_urls.append(full_url)

        # To‘g‘ri javobni aniqlash
        correct_answer = ""
        if question_type == 'text':
            correct_answer = re.sub(r'<[^>]+>', '', question.correct_text_answer_uz if lang == 'uz' else question.correct_text_answer_ru or "")
        elif question_type in ['choice', 'image_choice']:
            correct_choices = question.choices.filter(is_correct=True)
            correct_answer = "\n".join(
                f"{choice.letter}. {re.sub(r'<[^>]+>', '', choice.text or '')}" for choice in correct_choices
            )
        elif question_type == 'composite':
            sub_questions = question.sub_questions.all()
            correct_answer = "\n".join(
                f"{sq.text1 or ''} => {sq.correct_answer} {sq.text2 or ''}" for sq in sub_questions
            )
        else:
            return Response({'error': 'Noma’lum savol turi'}, status=status.HTTP_400_BAD_REQUEST)

        # Tilga qarab prompt yaratish
        if lang == 'uz':
            prompt_text = (
                f"Savol: {question_text}\n\n"
                f"Iltimos, yuqoridagi savolni tushunarli tarzda yechib bering, ishlanish yo‘li bilan. "
                f"Menda quyidagi to‘g‘ri javob mavjud — sizning yechimingiz ham aynan shunga to‘g‘ri kelishi kerak.\n\n"
                f"To‘g‘ri javob: {correct_answer}"
            )
        elif lang == 'ru':
            prompt_text = (
                f"Вопрос: {question_text}\n\n"
                f"Пожалуйста, решите приведённый выше вопрос с подробным объяснением и обоснованием решения. "
                f"Ваш ответ должен совпадать с правильным ответом, приведённым ниже.\n\n"
                f"Правильный ответ: {correct_answer}"
            )

        # OpenAI uchun xabar
        message_content = []

        for url in image_urls:
            message_content.append({
                "type": "image_url",
                "image_url": {
                    "url": url,
                    "detail": "auto"
                }
            })

        message_content.append({
            "type": "text",
            "text": prompt_text
        })

        # OpenAI bilan ishlash
        try:
            completion = client.chat.completions.create(
                model="gpt-4o",
                messages=[{
                    "role": "user",
                    "content": message_content
                }]
            )
        except Exception as e:
            return Response({'error': f'AI bilan bog‘lanishda xatolik: {str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        result_content = completion.choices[0].message.content

        return Response({
            'question': question_text,
            'images': image_urls,
            'correct_answer': correct_answer,
            'ai_response': result_content
        })